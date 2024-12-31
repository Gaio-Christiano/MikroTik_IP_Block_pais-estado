import re
import pandas as pd
from datetime import datetime
from time import sleep
import requests
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

# Caminho do arquivo exportado do Mikrotik
file_path = "portascan-list.txt"

# Verificar se o arquivo existe
if not os.path.exists(file_path):
    print(f"Erro: O arquivo '{file_path}' não foi encontrado.")
    exit()

# Expressão regular para extrair IP, Data, Hora e Timeout (Timeout é opcional)
pattern = r"PORTASCAN\s+([\d\.]+)\s+(\d{4}-\d{2}-\d{2})\s+([\d:]+)(?:\s+([\d\w]+))?"

# Lista para armazenar os dados
dados = []

# Ler o arquivo e validar as linhas
with open(file_path, "r") as arquivo:
    for linha in arquivo:
        match = re.search(pattern, linha)
        if match:
            ip = match.group(1)
            data = match.group(2)
            hora = match.group(3)
            timeout = match.group(4) if match.group(4) else "Sem Timeout"
            dados.append({"IP": ip, "Data": data, "Hora": hora, "Timeout": timeout})
        else:
            print(f"Linha ignorada (formato inválido): {linha.strip()}")

# Verificar se há dados processados
if not dados:
    print("Erro: Nenhum dado válido foi encontrado no arquivo.")
    exit()

# Funções para consultar múltiplas APIs
def consultar_geolocalizacao_api1(ip):
    try:
        response = requests.get(f"https://ipwhois.app/json/{ip}", timeout=10)
        if response.status_code == 200:
            data = response.json()
            return (
                data.get("city", "Desconhecida"),
                data.get("region", "Desconhecida"),
                data.get("country", "Desconhecido"),
                data.get("zip", "Desconhecido"),
            )
    except Exception as e:
        print(f"Erro na API 1 para {ip}: {e}")
    return "Desconhecida", "Desconhecida", "Desconhecido", "Desconhecido"

def consultar_geolocalizacao_api2(ip):
    try:
        access_key = "SUA_CHAVE_DE_API"
        response = requests.get(f"http://api.ipstack.com/{ip}?access_key={access_key}", timeout=10)
        if response.status_code == 200:
            data = response.json()
            return (
                data.get("city", "Desconhecida"),
                data.get("region_name", "Desconhecida"),
                data.get("country_name", "Desconhecido"),
                data.get("zip", "Desconhecido"),
            )
    except Exception as e:
        print(f"Erro na API 2 para {ip}: {e}")
    return "Desconhecida", "Desconhecida", "Desconhecido", "Desconhecido"

def consultar_geolocalizacao_api3(ip):
    try:
        response = requests.get(f"http://ip-api.com/json/{ip}?fields=city,regionName,country,zip", timeout=10)
        if response.status_code == 200:
            data = response.json()
            return (
                data.get("city", "Desconhecida"),
                data.get("regionName", "Desconhecida"),
                data.get("country", "Desconhecido"),
                data.get("zip", "Desconhecido"),
            )
    except Exception as e:
        print(f"Erro na API 3 para {ip}: {e}")
    return "Desconhecida", "Desconhecida", "Desconhecido", "Desconhecido"

def consultar_geolocalizacao(ip):
    for consulta in [consultar_geolocalizacao_api1, consultar_geolocalizacao_api2, consultar_geolocalizacao_api3]:
        cidade, estado, pais, cep = consulta(ip)
        if pais != "Desconhecido":
            return cidade, estado, pais, cep
    return "Desconhecida", "Desconhecida", "Desconhecido", "Desconhecido"

# Processar IPs e buscar geolocalização
for idx, entrada in enumerate(dados, start=1):
    ip = entrada["IP"]
    print(f"Processando IP {idx}/{len(dados)}: {ip}")
    cidade, estado, pais, cep = consultar_geolocalizacao(ip)
    entrada["Cidade"] = cidade
    entrada["Estado"] = estado
    entrada["País"] = pais
    entrada["CEP"] = cep
    entrada["Província"] = estado
    entrada["Bairro"] = cidade if cidade != "Desconhecida" else "Desconhecido"
    sleep(1)

# Criar um DataFrame do pandas
df = pd.DataFrame(dados)

# Resumo por País
resumo_pais = (
    df.groupby("País")
    .agg(
        Quantidade=("País", "size"),
        Estados_Uniquos=("Estado", lambda x: len(set(x))),
        Cidades_Uniquas=("Cidade", lambda x: len(set(x))),
    )
    .reset_index()
)

# Estados por País
estados_por_pais = (
    df.groupby(["País", "Estado"])
    .agg(Quantidade=("Estado", "size"))
    .reset_index()
)

# Bairros por Estado e País
bairros_por_estado_pais = (
    df.groupby(["País", "Estado", "Bairro"])
    .agg(Quantidade=("Bairro", "size"))
    .reset_index()
)

# Porcentagem por País
total_ips = len(df)
porcentagem_pais = (
    df.groupby("País")
    .agg(Quantidade=("País", "size"))
    .assign(Percentual=lambda x: (x["Quantidade"] / total_ips) * 100)
    .reset_index()
)

# Porcentagem por Estado
porcentagem_estado = (
    df.groupby(["País", "Estado"])
    .agg(Quantidade=("Estado", "size"))
    .assign(Percentual=lambda x: (x["Quantidade"] / total_ips) * 100)
    .reset_index()
)

# Resumo de localização
colunas = ["Cidade", "Estado", "País", "CEP"]
localizacao_resumo = {
    "Coluna": colunas,
    "Localizados (%)": [
        100 - ((df[coluna] == "Desconhecida").sum() / total_ips * 100) for coluna in colunas
    ],
}
df_localizacao_resumo = pd.DataFrame(localizacao_resumo)

# Nome do arquivo com dia e hora
hora_atual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
arquivo_saida = f"portascan_Relatorio_Completo_{hora_atual}.xlsx"

# Salvar como Excel com as abas adicionais e aplicar formatações
with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Detalhes", index=False)
    resumo_pais.to_excel(writer, sheet_name="Resumo por País", index=False)
    estados_por_pais.to_excel(writer, sheet_name="Estados por País", index=False)
    bairros_por_estado_pais.to_excel(writer, sheet_name="Bairros por Estado e País", index=False)
    porcentagem_pais.to_excel(writer, sheet_name="Porcentagem por País", index=False)
    porcentagem_estado.to_excel(writer, sheet_name="Porcentagem por Estado", index=False)
    df_localizacao_resumo.to_excel(writer, sheet_name="Resumo de Localização", index=False)

# Aplicar formatações na planilha
wb = load_workbook(arquivo_saida)

# Função para ajustar largura das colunas
def ajustar_largura_colunas(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 2

# Formatação para a aba "Detalhes"
ws = wb["Detalhes"]
ws.freeze_panes = "B2"  # Congela o painel
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal="center")  # Centraliza os dados

# Aplicar autofiltragem e formatação condicional
ws.auto_filter.ref = ws.dimensions

# Cabeçalhos com fundo azul e negrito
header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
header_font = Font(bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Ajustar largura das colunas na aba "Detalhes"
ajustar_largura_colunas(ws)

# Aplicar ajustes às demais abas
for aba in ["Resumo por País", "Estados por País", "Bairros por Estado e País", "Porcentagem por País", "Porcentagem por Estado", "Resumo de Localização"]:
    ws = wb[aba]
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ajustar_largura_colunas(ws)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

# Salvar o arquivo formatado
wb.save(arquivo_saida)

print(f"Arquivo salvo e formatado em: {arquivo_saida}")